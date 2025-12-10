# ============================================================================
# FILE: finanalyst_pro/tools/ingestion/excel_parser.py
# PURPOSE: Parse financial statements from Excel files
# ============================================================================

from __future__ import annotations

import time
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any

from pydantic import BaseModel, Field
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from finanalyst_pro.core.infrastructure import (
    ToolResult, AuditLogEntry, AUDIT, ConfidenceLevel
)
from finanalyst_pro.schemas.financial_statements import (
    FinancialStatements, IncomeStatementLine, BalanceSheetLine, 
    CashFlowLine, Period
)


# ============================================================================
# INPUT/OUTPUT SCHEMAS
# ============================================================================

class SheetMapping(BaseModel):
    """Configuration for mapping Excel sheets to statement types."""
    
    income_statement_sheet: str | None = None
    balance_sheet_sheet: str | None = None
    cash_flow_sheet: str | None = None
    
    # Row/column hints for data extraction
    header_row: int = 1
    data_start_row: int = 2
    label_column: int = 1  # Column A = 1
    
    # Year columns (e.g., {2020: 2, 2021: 3, 2022: 4})
    year_columns: dict[int, int] = Field(default_factory=dict)


class ExcelParseInput(BaseModel):
    """Input schema for Excel parsing tool."""
    
    file_path: str = Field(
        description="Path to the Excel file containing financial statements"
    )
    sheet_mapping: SheetMapping | None = Field(
        default=None,
        description="Optional mapping configuration for sheet names and columns"
    )
    auto_detect: bool = Field(
        default=True,
        description="Attempt to auto-detect sheet structure if mapping not provided"
    )
    company_name: str | None = Field(
        default=None,
        description="Company name to include in parsed output"
    )
    currency: str = Field(
        default="SGD",
        description="Currency code for monetary values"
    )


class ExcelParseOutput(BaseModel):
    """Output schema for Excel parsing tool."""
    
    financial_statements: FinancialStatements
    parsing_confidence: ConfidenceLevel
    sheets_found: list[str]
    sheets_parsed: list[str]
    warnings: list[str] = Field(default_factory=list)
    field_coverage: dict[str, float] = Field(
        default_factory=dict,
        description="Percentage of expected fields found per statement type"
    )


# ============================================================================
# TOOL IMPLEMENTATION
# ============================================================================

class ExcelFinancialParser:
    """
    Tool for parsing financial statements from Excel files.
    
    Handles common formats including:
    - Multi-sheet workbooks with separate IS/BS/CF tabs
    - Single-sheet consolidated statements
    - Various row/column orientations
    """
    
    TOOL_NAME = "parse_excel_financial_statements"
    
    # Common sheet name patterns for auto-detection
    INCOME_STATEMENT_PATTERNS = [
        "income statement", "income", "p&l", "profit and loss", 
        "profit & loss", "operations", "statement of operations"
    ]
    BALANCE_SHEET_PATTERNS = [
        "balance sheet", "balance", "bs", "financial position",
        "statement of financial position"
    ]
    CASH_FLOW_PATTERNS = [
        "cash flow", "cash flows", "cf", "statement of cash flows"
    ]
    
    # Common row labels for auto-detection
    REVENUE_LABELS = ["revenue", "total revenue", "net sales", "total net sales", "sales"]
    COGS_LABELS = ["cost of sales", "cost of goods sold", "cogs", "cost of revenue"]
    NET_INCOME_LABELS = ["net income", "net profit", "net earnings", "profit for the year"]
    
    def execute(self, input_data: ExcelParseInput) -> ToolResult[ExcelParseOutput]:
        """
        Parse financial statements from an Excel file.
        
        Args:
            input_data: Parsing configuration and file path
            
        Returns:
            ToolResult containing parsed financial statements or error
        """
        start_time = time.time()
        warnings: list[str] = []
        
        try:
            # Validate file exists
            file_path = Path(input_data.file_path)
            if not file_path.exists():
                return self._error_result(f"File not found: {file_path}", start_time)
            
            if not file_path.suffix.lower() in [".xlsx", ".xls", ".xlsm"]:
                return self._error_result(
                    f"Unsupported file format: {file_path.suffix}", start_time
                )
            
            # Load workbook
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheets_found = workbook.sheetnames
            
            # Determine sheet mapping
            if input_data.sheet_mapping:
                mapping = input_data.sheet_mapping
            elif input_data.auto_detect:
                mapping, detect_warnings = self._auto_detect_mapping(workbook)
                warnings.extend(detect_warnings)
            else:
                return self._error_result(
                    "No sheet mapping provided and auto_detect is disabled", 
                    start_time
                )
            
            # Parse each statement type
            income_statements = []
            balance_sheets = []
            cash_flows = []
            sheets_parsed = []
            field_coverage = {}
            
            # Parse Income Statement
            if mapping.income_statement_sheet:
                if mapping.income_statement_sheet in sheets_found:
                    sheet = workbook[mapping.income_statement_sheet]
                    is_data, is_warnings, is_coverage = self._parse_income_statement(
                        sheet, mapping
                    )
                    income_statements.extend(is_data)
                    warnings.extend(is_warnings)
                    sheets_parsed.append(mapping.income_statement_sheet)
                    field_coverage["income_statement"] = is_coverage
                else:
                    warnings.append(
                        f"Income statement sheet '{mapping.income_statement_sheet}' not found"
                    )
            
            # Parse Balance Sheet
            if mapping.balance_sheet_sheet:
                if mapping.balance_sheet_sheet in sheets_found:
                    sheet = workbook[mapping.balance_sheet_sheet]
                    bs_data, bs_warnings, bs_coverage = self._parse_balance_sheet(
                        sheet, mapping
                    )
                    balance_sheets.extend(bs_data)
                    warnings.extend(bs_warnings)
                    sheets_parsed.append(mapping.balance_sheet_sheet)
                    field_coverage["balance_sheet"] = bs_coverage
                else:
                    warnings.append(
                        f"Balance sheet '{mapping.balance_sheet_sheet}' not found"
                    )
            
            # Parse Cash Flow
            if mapping.cash_flow_sheet:
                if mapping.cash_flow_sheet in sheets_found:
                    sheet = workbook[mapping.cash_flow_sheet]
                    cf_data, cf_warnings, cf_coverage = self._parse_cash_flow(
                        sheet, mapping
                    )
                    cash_flows.extend(cf_data)
                    warnings.extend(cf_warnings)
                    sheets_parsed.append(mapping.cash_flow_sheet)
                    field_coverage["cash_flow"] = cf_coverage
                else:
                    warnings.append(
                        f"Cash flow sheet '{mapping.cash_flow_sheet}' not found"
                    )
            
            # Build financial statements object
            financial_statements = FinancialStatements(
                company_name=input_data.company_name,
                currency=input_data.currency,
                income_statements=income_statements,
                balance_sheets=balance_sheets,
                cash_flow_statements=cash_flows,
                source_file=str(file_path),
                extraction_timestamp=time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
            )
            
            # Determine parsing confidence
            confidence = self._assess_confidence(
                income_statements, balance_sheets, cash_flows, field_coverage
            )
            
            output = ExcelParseOutput(
                financial_statements=financial_statements,
                parsing_confidence=confidence,
                sheets_found=sheets_found,
                sheets_parsed=sheets_parsed,
                warnings=warnings,
                field_coverage=field_coverage,
            )
            
            # Log audit entry
            self._log_audit(input_data, True, start_time)
            
            execution_time = (time.time() - start_time) * 1000
            
            return ToolResult(
                success=True,
                tool_name=self.TOOL_NAME,
                data=output.model_dump(),
                warnings=warnings,
                execution_time_ms=execution_time,
                metadata={
                    "file": str(file_path),
                    "sheets_parsed": len(sheets_parsed),
                    "periods_found": len(financial_statements.get_periods()),
                }
            )
            
        except Exception as e:
            self._log_audit(input_data, False, start_time, str(e))
            return self._error_result(str(e), start_time)
    
    def _auto_detect_mapping(
        self, 
        workbook: openpyxl.Workbook
    ) -> tuple[SheetMapping, list[str]]:
        """Auto-detect sheet mapping from workbook structure."""
        warnings = []
        mapping = SheetMapping()
        
        for sheet_name in workbook.sheetnames:
            name_lower = sheet_name.lower().strip()
            
            if any(p in name_lower for p in self.INCOME_STATEMENT_PATTERNS):
                mapping.income_statement_sheet = sheet_name
            elif any(p in name_lower for p in self.BALANCE_SHEET_PATTERNS):
                mapping.balance_sheet_sheet = sheet_name
            elif any(p in name_lower for p in self.CASH_FLOW_PATTERNS):
                mapping.cash_flow_sheet = sheet_name
        
        # Try to detect year columns from first detected sheet
        detected_sheet = (
            mapping.income_statement_sheet or 
            mapping.balance_sheet_sheet or 
            mapping.cash_flow_sheet
        )
        
        if detected_sheet:
            sheet = workbook[detected_sheet]
            year_columns = self._detect_year_columns(sheet)
            mapping.year_columns = year_columns
        
        if not any([
            mapping.income_statement_sheet,
            mapping.balance_sheet_sheet,
            mapping.cash_flow_sheet
        ]):
            warnings.append(
                "Could not auto-detect any financial statement sheets. "
                "Please provide explicit sheet_mapping."
            )
        
        return mapping, warnings
    
    def _detect_year_columns(self, sheet: Worksheet) -> dict[int, int]:
        """Detect which columns contain which fiscal years."""
        year_columns = {}
        
        # Check first few rows for year headers
        for row_idx in range(1, 5):
            for col_idx in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    # Try to extract year from cell
                    year = self._extract_year(cell_value)
                    if year and 1990 <= year <= 2100:
                        year_columns[year] = col_idx
        
        return year_columns
    
    def _extract_year(self, value: Any) -> int | None:
        """Extract a year from a cell value."""
        if isinstance(value, int) and 1990 <= value <= 2100:
            return value
        if isinstance(value, str):
            # Try patterns like "2022", "FY2022", "FY 2022"
            import re
            match = re.search(r'\b(19|20)\d{2}\b', value)
            if match:
                return int(match.group())
        return None
    
    def _parse_income_statement(
        self,
        sheet: Worksheet,
        mapping: SheetMapping
    ) -> tuple[list[IncomeStatementLine], list[str], float]:
        """Parse income statement data from a sheet."""
        statements = []
        warnings = []
        fields_found = 0
        fields_expected = 10  # Key fields we expect
        
        # Build label-to-row mapping
        label_rows = {}
        for row_idx in range(mapping.data_start_row, sheet.max_row + 1):
            label = sheet.cell(row=row_idx, column=mapping.label_column).value
            if label:
                label_rows[str(label).lower().strip()] = row_idx
        
        # Parse each year column
        for year, col_idx in mapping.year_columns.items():
            period = Period(
                start_date=date(year, 1, 1),
                end_date=date(year, 12, 31),
                fiscal_year=year,
                period_type="annual"
            )
            
            # Extract values by matching labels
            stmt_data = {"period": period}
            
            # Revenue
            revenue_row = self._find_row(label_rows, self.REVENUE_LABELS)
            
